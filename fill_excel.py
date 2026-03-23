#!/usr/bin/env python3
"""
動支及黏存單填寫腳本
使用 openpyxl 填入資料，完整保留原始格式（欄寬、列高、框線、合併儲存格）

用法（由 run.py 呼叫，不需直接執行）：
  python fill_excel.py --data '{"items":[...],...}' --output output.xlsx
"""

import json
import argparse
import sys
import os
from copy import copy
import openpyxl
from openpyxl import load_workbook


# 品項所在列號（1-indexed）
ITEM_ROWS = [15, 17, 19, 21, 23, 25, 27, 29]

# 各 sheet 的欄位對應
SHEET_COLUMNS = {
    '預算內': {
        'name': 'A',      # 名稱及規格
        'qty':  'C',      # 數量
        'price': 'E',     # 單價
        'purpose': 'I',   # 用途說明（第15列）
        'unit': 'B',      # 單位別（B13）
        'month_col': 'I', # 月（I13）
        'day_col': 'K',   # 日（K13）
        'cat_row': 4,     # 一級科目列
        'sub_row': 5,     # 二級科目列
        'cat_col': 'B',
    },
    '代收代辦': {
        'name': 'A',
        'qty':  'C',
        'price': 'E',
        'purpose': 'J',   # 用途說明（第15列）
        'unit': 'B',
        'month_col': 'J', # 月（J13）
        'day_col': 'L',   # 日（L13）
        'cat_row': 4,
        'sub_row': 5,
        'cat_col': 'B',
    }
}


def col_letter_to_idx(letter):
    """欄位字母轉 0-indexed 數字"""
    result = 0
    for c in letter:
        result = result * 26 + (ord(c.upper()) - ord('A') + 1)
    return result - 1


def set_cell(ws, col_letter, row, value):
    """安全設定儲存格值，保留原有格式"""
    cell = ws[f'{col_letter}{row}']
    # 處理合併儲存格：找到主儲存格
    cell.value = value


def clear_item_rows(ws, cols):
    """清除所有品項列的舊資料（只清名稱、數量、單價）"""
    for row in ITEM_ROWS:
        for col in [cols['name'], cols['qty'], cols['price']]:
            ws[f'{col}{row}'].value = None


def fill_data(ws, data, cols):
    """填入品項資料"""
    items = data.get('items', [])[:8]  # 最多 8 筆

    # 清除舊資料
    clear_item_rows(ws, cols)

    # 填入新資料
    for i, item in enumerate(items):
        row = ITEM_ROWS[i]
        name = item.get('name', '')
        qty = item.get('quantity', 0)
        price = item.get('unitPrice', 0)

        if name:
            ws[f"{cols['name']}{row}"].value = name
        if qty:
            ws[f"{cols['qty']}{row}"].value = qty
        if price:
            ws[f"{cols['price']}{row}"].value = price

    # 所屬年度（B2）
    year = data.get('year')
    if year:
        ws['B2'].value = int(year)

    # 用途說明（第 15 列）
    purpose = data.get('purpose', '')
    if purpose:
        ws[f"{cols['purpose']}15"].value = purpose

    # 單位別（B13）
    unit = data.get('unit', '')
    if unit:
        ws[f"{cols['unit']}13"].value = unit

    # 月、日（年由公式自動帶入）
    month = data.get('month')
    day = data.get('day')
    if month:
        ws[f"{cols['month_col']}13"].value = int(month)
    if day:
        ws[f"{cols['day_col']}13"].value = int(day)

    # 預算科目
    cat = data.get('budgetCategory', '')
    sub = data.get('budgetSubCategory', '')
    if cat:
        ws[f"{cols['cat_col']}{cols['cat_row']}"].value = cat
    if sub:
        ws[f"{cols['cat_col']}{cols['sub_row']}"].value = sub


def process(input_path, data, output_path):
    """主處理函式"""
    wb = load_workbook(input_path)

    sheet_name = data.get('templateType', '預算內')
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'找不到工作表「{sheet_name}」')

    ws = wb[sheet_name]
    cols = SHEET_COLUMNS[sheet_name]

    fill_data(ws, data, cols)

    # 設定開啟時顯示正確的工作表，並清除另一張的舊資料
    wb.active = wb[sheet_name]
    other_name = '代收代辦' if sheet_name == '預算內' else '預算內'
    if other_name in wb.sheetnames:
        ws_other = wb[other_name]
        other_cols = SHEET_COLUMNS[other_name]
        clear_item_rows(ws_other, other_cols)

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description='填寫動支及黏存單')
    parser.add_argument('--data', required=True, help='JSON 資料字串')
    parser.add_argument('--template', default='template/template.xlsx', help='範本路徑')
    parser.add_argument('--output', required=True, help='輸出路徑')
    args = parser.parse_args()

    data = json.loads(args.data)
    process(args.template, data, args.output)
    print(args.output)


if __name__ == '__main__':
    main()
